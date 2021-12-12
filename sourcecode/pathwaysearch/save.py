# -*- coding: utf-8 -*-
"""
Created on Wed Nov 17 11:58:38 2021

@author: CC-SXF
"""

import os
import json
from openpyxl import Workbook
from openpyxl.styles import Font, colors
from openpyxl.utils import get_column_letter

class Save():
    """
    """

    def __init__(self,):
        """ """
        pass

    @classmethod
    def _createSheet(cls, workbook, sheet_name, data_list=[], highlight_cell_list=[]):
        """
        """
        # Create sheet "sheet_name"
        worksheet = workbook.create_sheet(sheet_name)
        worksheet.freeze_panes = "B2"
        # Write title and entries
        for entry in data_list:
            try:
                worksheet.append(entry)
            except Exception as e:
                print(e)
                print(entry)
                print("")
        # Set the format of the sheet
        font_format = Font(name='Times New Roman', bold=True, color=colors.BLACK)
        for idx_col in range(len(data_list[0])):
            col_letter = get_column_letter(idx_col+1)
            worksheet.column_dimensions[col_letter].width = 15
            cell = col_letter + "1"
            worksheet[cell].font = font_format
        # worksheet.column_dimensions['A'].width = 10
        font_format = Font(name='Times New Roman', bold=True, color=colors.BLUE)
        for cell in highlight_cell_list:
            worksheet[cell].font = font_format
        return True

    @classmethod
    def save(cls, pathways):
        """
        """
        dir_details = "./results/pathways/"
        os.makedirs(dir_details)
        dir_pathways = os.path.join(dir_details, "pathways.xlsx")
        # Create workbook
        workbook = Workbook()
        workbook.guess_types = True
        if "Sheet" in workbook.sheetnames:
            workbook.remove(workbook["Sheet"])

        title = ["Feasibility",
                 "Total-Length", "Endo-Length", "Heter-Length", "Inf-Length",
                 "Atom-Utilization", "Atom-Conservation", "Metabolic-Flux",
                 "S-Details", "I-Details", "M-Details", "F-Details"]
        data_list = list()
        data_list.append(title)
        for entry in pathways:
            for idx, item in enumerate(entry):
                if type(item) == list:
                    entry[idx] = json.dumps(item)
                else:
                    entry[idx] = item
            data_list.append(entry)

        cls._createSheet(workbook, "pathways", data_list, ['B1', 'I1'])

        # Save the excel file
        workbook.save(dir_pathways)


if __name__ == "__main__":
    """ """
    pass





