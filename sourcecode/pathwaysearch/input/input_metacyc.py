# -*- coding: utf-8 -*-
"""
Created on Thu Dec  3 22:22:30 2020

@author: CC-SXF
"""

import re
import json
# import copy
import pickle
from os import path, listdir, mkdir

from openpyxl import Workbook
# from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, colors
# from openpyxl.styles import PatternFill

from rdkit import Chem

from pathwaysearch.input.input_meta import InputMeta
from pathwaysearch.rxndir import RxnDir
from pathwaysearch.reapropair import ReaProPair
from pathwaysearch.reconcile import Reconcile
from pathwaysearch.fingerprint.fingerprint import FingerPrint

from pathwaysearch.ltiod import Ltiod


class InputMetaCyc():
    """
    """
    def __init__(self,):
        """ """
        pass


    @classmethod
    def _createSheet(cls, workbook, sheet_name, data_list=[], highlight_cell_list=[]):
        """
        """
        # Create sheet "sheet_name"
        worksheet = workbook.create_sheet(sheet_name)
        worksheet.freeze_panes = "B2"
        # Write title and entries
        for comp_info in data_list:
            worksheet.append(comp_info)
        # Set the format of the sheet
        font_format = Font(name='Times New Roman', bold=True, color=colors.BLACK)
        for idx_col in range(len(data_list[0])):
            col_letter = get_column_letter(idx_col+1)
            worksheet.column_dimensions[col_letter].width = 15
            cell = col_letter + "1"
            worksheet[cell].font = font_format
        worksheet.column_dimensions['A'].width = 10
        font_format = Font(name='Times New Roman', bold=True, color=colors.BLUE)
        for cell in highlight_cell_list:
            worksheet[cell].font = font_format


    @classmethod
    def _getMetaCycCompRxnDistInfo(cls, dirSourceData = "../sourcedata/",
                                        resource_xlsxFile = "KndPad_Ori_Cleaned_Half.xlsx",
                                        metacyc_xref_kndpad_jsonFile = "MetaCyc_Xref_KndPad.json",
                                        ):
        """
            Distill the information from metaCyc.
        """
        resource_xlsxFile = path.join(dirSourceData, resource_xlsxFile)
        compInfoList = InputMeta._getSheetInfo(resource_xlsxFile, sheetName="Compound")
        rxnInfoList = InputMeta._getSheetInfo(resource_xlsxFile, sheetName="InfoEx")
        #
        metacyc_xref_kndpad_dict = dict()
        metacyc_xref_kndpad_dict["Compound"] = dict()
        metacyc_xref_kndpad_dict["Reaction"] = dict()
        #
        compInPsRxnSet = set()
        rxnInPsSet = set()
        rxnDistInfoList = list()
        title = rxnInfoList[0]
        # Entry(0); Names(1); Definition(2); Equation(3); Coefficient(4); Compound(5); SMILES(6);
        # SMILES Equation; Atom/Bond Number;   --> delete
        # Delta Gibbs(kcal/mol)(8); EC Number(9);
        # Enzymes; Pathways;      --> delete
        # Status(7);
        #
        # Source(10); MetaCyc(11); Rhea(12); KEGG(13); BiGG(14); MetaNetX(15); UniProt(16); Reference(17);
        #
        #                    Status;        Delta Gibbs(kcal/mol); EC Number;
        title = title[:7] + title[13:14] + title[9:11] + title[14:]
        rxnDistInfoList.append(title)
        for rxnInfo in rxnInfoList[1:]:  # Exclude title
            rxnId = rxnInfo[0]                                          # Entry
            if not rxnId.endswith('-m'):
                continue
            rxnEquation = rxnInfo[3]                                    # Equation
            rxnCompList = re.findall("Met\d{6}-m", rxnEquation)
            compInPsRxnSet |= set(rxnCompList)
            rxnInPsSet.add(rxnId)
            #
            source = rxnInfo[14]                                        # Source
            source = source.replace("MetaCyc:", "")
            source = source.replace("%2b", "+")
            metacyc_xref_kndpad_dict["Reaction"][source] = rxnId
            #
            rxnDistInfo = rxnInfo[:7] + rxnInfo[13:14] + rxnInfo[9:11] + rxnInfo[14:]
            rxnDistInfoList.append(rxnDistInfo)
        #
        #
        compDistInfoList = list()
        title = compInfoList[0]
        # Entry(0); Names(1); Chemical Formula(2);
        # Molecular Weight(Daltons)(4); SMILES(5);
        # Delta Gibbs(kcal/mol)(3);
        # Reaction Number;                               --> delete
        # Reactions(6);
        #
        # Source(7); MetaCyc(8); ChEBI(9); KEGG(10); BiGG(11); MetaNetX(12); PubChem(13); Reference(14);
        #                    Delta Gibbs(kcal/mol);    Reactions
        title = title[:3] + title[5:6] + title[3:5] + title[7:8] +  title[8:]
        compDistInfoList.append(title)
        for idx, compInfo in enumerate(compInfoList[1:]):  # Exclude title
            compId = compInfo[0]                                        # Entry
            if not compId.endswith('-m'):
                continue
            if compId in compInPsRxnSet:
                compRxns = compInfo[7]                                  # Reactions
                compRxns = json.loads(compRxns)
                compRxns = set(compRxns)
                compRxns = compRxns & rxnInPsSet
                compRxns = sorted(compRxns)
                compRxns = json.dumps(compRxns)
                compInfo[7] = compRxns
                #
                source = compInfo[8]                                    # Source
                source = source.replace("MetaCyc:", "")
                source = source.replace("%2b", "+")
                metacyc_xref_kndpad_dict["Compound"][source] = compId

                #
                compDistInfo = compInfo[:3] + compInfo[5:6] + compInfo[3:5] + compInfo[7:8] +  compInfo[8:]
                compDistInfoList.append(compDistInfo)
        #
        # Saving...
        if "datas" not in listdir():
            mkdir("datas/")
        metacyc_xref_kndpad_jsonFile = path.join('datas', metacyc_xref_kndpad_jsonFile)
        with open(metacyc_xref_kndpad_jsonFile, "w") as file_obj:
            file_obj.write(json.dumps(metacyc_xref_kndpad_dict, indent=2))
        #
        return [compDistInfoList, rxnDistInfoList]


    @classmethod
    def _getMetaCycRxnInfo(cls, dirSourceData = "../sourcedata/",
                                resource_xlsxFile = "KndPad_Ori_Cleaned_Half.xlsx",
                                ):
        """
            Get info of every reaction's coef/coefSum/compID/canoSmiles/nonIsoSmiles and
        isoReaProReactionIDs from the cleaned MetaCyc(KndPad) reaction excle(InfoEx).
            rxnIsoReaProIdSet: Reaction IDs set of which the Reactants or Products have
        Isomeric Compounds.
        """
        resource_xlsxFile = path.join(dirSourceData, resource_xlsxFile)
        #
        InfoEx = InputMeta._getSheetInfo(resource_xlsxFile, sheetName = "InfoEx")
        #
        reactionCoefCompSmiDict = dict()
        # Reaction IDs set of which the Reactants or Products have Isomeric Compounds.
        reactionIsoReaProIdSet = set()
        for row in InfoEx[1:]:    # Exclude Title
            reactionID = row[0]
            if not reactionID.endswith('-m'):
                continue
            # Parse Coefficient
            reactionCoefStr = row[4]
            reactionCoefList = json.loads(reactionCoefStr)
            reactionCoefSumList =[sum(reactionCoefList[0]), sum(reactionCoefList[1])]
            # Parse Compound
            reactionCompStr = row[5]
            reactionCompList = json.loads(reactionCompStr)
            # Parse Canonical Smiles
            reactionCanoSmiStr = row[6]
            reactionCanoSmiList = json.loads(reactionCanoSmiStr)
            # Non Isomeric Smiles
            reactionNonIsoSmiList = [[],[]]
            for canoSmi in reactionCanoSmiList[0]:
                reactionNonIsoSmiList[0].append( Chem.MolToSmiles(Chem.MolFromSmiles(canoSmi), isomericSmiles=False) )
            for canoSmi in reactionCanoSmiList[1]:
                reactionNonIsoSmiList[1].append( Chem.MolToSmiles(Chem.MolFromSmiles(canoSmi), isomericSmiles=False) )
            # Isomeric Compound In Reactants Or Products. Replace Non-Isomeric Smiles with Isomeric Smiles.
            # We will employ the same processing strategy in further process.
            if (len(set(reactionCanoSmiList[0])) != len(set(reactionNonIsoSmiList[0]))):
                print(reactionID,": Isomeric Compounds In Reactants.")
                reactionNonIsoSmiList[0] = reactionCanoSmiList[0][:]
                reactionNonIsoSmiList[1] = reactionCanoSmiList[1][:]
                reactionIsoReaProIdSet.add(reactionID)
            elif (len(set(reactionCanoSmiList[1])) != len(set(reactionNonIsoSmiList[1]))):
                print(reactionID,": Isomeric Compounds In Products.")
                reactionNonIsoSmiList[0] = reactionCanoSmiList[0][:]
                reactionNonIsoSmiList[1] = reactionCanoSmiList[1][:]
                reactionIsoReaProIdSet.add(reactionID)
            else:
                pass
            reactionCoefCompSmiDict[reactionID] = [reactionCoefList, reactionCoefSumList,
                                                   reactionCompList,
                                                   reactionCanoSmiList, reactionNonIsoSmiList]
        return [reactionCoefCompSmiDict, reactionIsoReaProIdSet]


    @classmethod
    def _getMetaCycGeCosInfo(cls, dirDatas ="datas/",
                                  general_cofactors_xlsxFile = "MetaCyc_General_Cofactors.xlsx",
                                  ):
        """ """
        general_cofactors_xlsxFile = path.join(dirDatas, general_cofactors_xlsxFile)
        sdqFamilyList = InputMeta._getGeCosInfoFromGeCosExcel(general_cofactors_xlsxFile, sheetName = "GeCos-New")
        return sdqFamilyList


    @classmethod
    def _getMetaCycPsRxnDirInfo(cls, ):
        """ """
        rxnDirInfoDict = dict()
        return rxnDirInfoDict


    @classmethod
    def _getMetaCycPsRxnCompDrPairInfo(cls, dirSourceData = "../sourcedata/",
                                            resource_xlsxFile = "KndPad_Ori_Cleaned_Half.xlsx",
                                            dirDatas = "datas/",
                                            general_cofactors_xlsxFile = "MetaCyc_General_Cofactors.xlsx",
                                            dirRxnFile = "../KndPadAAMBCsRCsFiles_Ori_Half/",
                                            diffThresholdLevel = 10,
                                            isDebug = True
                                            ):
        """
            Get the compound pairs of every reaction.
        """
        #
        with open(path.join("temp", "MetaCyc_Info.data"), "rb") as file_obj:
            [rxnCoefCompSmiDict, rxnIsoReaProIdSet] = pickle.load(file_obj)
        #
        """
        [rxnCoefCompSmiDict, rxnIsoReaProIdSet] = cls._getMetaCycRxnInfo(dirSourceData = dirSourceData,
                                                                         resource_xlsxFile = resource_xlsxFile,
                                                                         )
        """
        #
        sdqFamilyList = cls._getMetaCycGeCosInfo(dirDatas = dirDatas,
                                                 general_cofactors_xlsxFile = general_cofactors_xlsxFile,
                                                 )
        #
        exSubStrTupleList = FingerPrint.getExSubStrInfo(dataBase = "MetaCyc")
        #
        rxnDrCompPairsInfoDict = dict()
        rxnInPsList = list(rxnCoefCompSmiDict.keys())
        #
        for rxnId in rxnInPsList:
            #
            rxnFile = rxnId + "_ECBLAST_smiles_AAM.rxn"
            rxnFile = path.join(dirRxnFile, rxnFile)
            rxnCoefCompSmi = rxnCoefCompSmiDict[rxnId]
            try:
                # Data format:
                # [comp_molBlock_list, comp_addHsMapNumSmiles_list, comp_id_list]
                rxnInputInfo = InputMeta._rxnFile2MolSmiComp(rxnId, rxnCoefCompSmi, rxnIsoReaProIdSet, rxnFile)
                if isDebug:
                    print(rxnId)
                rxnCompMolBlockList = rxnInputInfo[0]
                rxnAddHsMnSmilesList = rxnInputInfo[1]
                rxnCompIdList = rxnInputInfo[2]
                #
                usaRxnCompIdList = InputMeta._getCompCands(rxnCompIdList, sdqFamilyList)
                #
                rxnDrCompPairsInfo = ReaProPair.getRxnCompDrPairs(usaRxnCompIdList, exSubStrTupleList,
                                                                  rxnCompMolBlockList, rxnAddHsMnSmilesList,
                                                                  diffThresholdLevel)
                rxnDrCompPairsInfoDict[rxnId] = rxnDrCompPairsInfo
                #
                if isDebug:
                    print("Compound Pairs")
                    diffThresholdScope = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0]
                    for diffThreshold, rxnCompDrPairs in zip(diffThresholdScope, rxnDrCompPairsInfo):
                        print(str(diffThreshold), ":", rxnCompDrPairs)
                    print("\n\n")
            except FileNotFoundError:
                pass
            except Exception as e:
                print(rxnId, ':', e, end='\n\n')
        #
        return rxnDrCompPairsInfoDict


    @classmethod
    def _fuseMetaCycPsRxnCompPairInfo(cls, rxnDistInfoList, rxnDirInfoDict, rxnDrCompPairsInfoDict,
                                           diffThresholdLevel = 10,
                                           ):
        """
        """
        diffThresholdScope = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0]
        diffThresholdList = diffThresholdScope[:diffThresholdLevel]
        compPairTitleList = list()
        for diffThreshold in diffThresholdList:
            compPairTitle = "".join(["Compound Pair ", "(", str(diffThreshold), ")"])
            compPairTitleList.append(compPairTitle)
        #
        # Entry; Names; Definition; Equation;
        # Direction;
        # Coefficient; Compound; Canonical Smiles;
        # Compound Pair (0.1); Compound Pair (0.2); Compound Pair (0.3); Compound Pair (0.4);
        # Compound Pair (0.5); Compound Pair (0.6); Compound Pair (0.7); Compound Pair (0.8);
        # Compound Pair (0.9);  Compound Pair (1.0);
        # Status; Delta Gibbs(kcal/mol); EC Number;
        # Source; MetaCyc; Rhea; KEGG; BiGG; MetaNetX; UniProt; Reference;
        title = rxnDistInfoList[0]
        title = title[:4] + ["Direction"] + title[4:7] + compPairTitleList + title[7:]
        #
        rxnFuseInfoList = list()
        rxnFuseInfoList.append(title)
        #
        # Some reactions' compound pair are manually reconciled.
        # These reactions are dropped in atom-atom mapping number computation because of various reasons.
        # Including 1) the valid maximum value of atom mapping number of '.mol' file is too small,
        # 2) identical structure of reactants and products,
        # 3) computing time is extremely long,
        # 4) the reaction direaction is obvious error(R03906),
        # 5) atom-mapping number file isn't generated by rdt.
        # The total number is 127.
        # We can choose to discard these reactions directly by reset the dictionary variable.
        # rxnDirCompPairDict = dict()
        rxnDirCompPairDict = Reconcile.reconcileMetaCyc()
        #
        for rxnDistInfo in rxnDistInfoList[1:]:  # Exclude title
            rxnId = rxnDistInfo[0]                              # Entry
            rxnEquation = rxnDistInfo[3]                        # Equation
            #
            # Direction of this reaction
            # rxnDir = rxnDirInfoDict[rxnId]
            rxnDir = RxnDir._getRxnDirection(rxnEquation)
            #
            try:
                rxnDrCompPairsInfo = rxnDrCompPairsInfoDict[rxnId]
            except KeyError:
                # null in .xlsx file
                rxnDrCompPairsInfo = [None] * diffThresholdLevel
            #
            # Reconciling...
            try:
                rxnDirCompPairList = rxnDirCompPairDict[rxnId]
                rxnDir = rxnDirCompPairList[0]
                rxnCompPair = json.loads(rxnDirCompPairList[1])
                rxnDrCompPairsInfo = [rxnCompPair] * diffThresholdLevel
            except KeyError:
                pass
            #
            for idx in range(len(rxnDrCompPairsInfo)):
                rxnDrCompPairsInfo[idx] = json.dumps(rxnDrCompPairsInfo[idx])
            #
            # Entry; Names; Definition; Equation;
            # Direction;
            # Coefficient; Compound; Canonical Smiles;
            # Compound Pair(0.1, 0.2, ..., 1.0);
            # Status; Delta Gibbs(kcal/mol); EC Number;
            # Source; MetaCyc; Rhea; KEGG; BiGG; MetaNetX; UniProt; Reference;
            rxnFuseInfo = rxnDistInfo[:4] + [rxnDir] + rxnDistInfo[4:7] + rxnDrCompPairsInfo + rxnDistInfo[7:]
            rxnFuseInfoList.append(rxnFuseInfo)
        return rxnFuseInfoList


    @classmethod
    def _gfLtiod(cls, compDistInfoList, rxnFuseInfoList,
                     diffThresholdLevel = 10, step = 2,
                     ):
        """
            Get and fuse the local total in/out-degree of every compound.
        """
        compFuseInfoList = Ltiod._gfLtiod(compDistInfoList, rxnFuseInfoList,
                                        diffThresholdLevel, step)
        return compFuseInfoList


    @classmethod
    def _saveMetaCycPsRxnCompPairInfo(cls, compFuseInfoList, rxnFuseInfoList,
                                           dirSourceData = "../sourcedata/",
                                           xlsxFile = "MetaCyc_Pathway_Search_Ori.xlsx",
                                           diffThresholdLevel = 10,
                                           ):
        """ """
        xlsxFile = path.join(dirSourceData, xlsxFile)
        #
        # Create workbook
        wb_rxncomp = Workbook()
        wb_rxncomp.guess_types = True
        if "Sheet" in wb_rxncomp.sheetnames:
            wb_rxncomp.remove(wb_rxncomp["Sheet"])
        #
        # Create sheet "Reaction"
        ws_rxn = wb_rxncomp.create_sheet("Reaction")
        ws_rxn.freeze_panes = "B2"
        # Write title and entries
        for rxnInfo in rxnFuseInfoList:
            ws_rxn.append(rxnInfo)
        # Set the format of the sheet
        #
        # Entry; Names; Definition; Equation; Direction;
        for col_letter in ["A", "B", "C", "D", "E"]:
            ws_rxn.column_dimensions[col_letter].width = 15
        #
        # Status; Delta Gibbs(kcal/mol); EC Number;
        # Source; MetaCyc; Rhea; KEGG; BiGG; MetaNetX; UniProt; Reference;
        for idx in range(diffThresholdLevel+9, diffThresholdLevel+20):
            col_letter = get_column_letter(idx)
            ws_rxn.column_dimensions[col_letter].width = 15
        #
        # Coefficient; Compound; Canonical Smiles; Compound Pair (0.1 ~ 1.0);
        for idx in range(6, diffThresholdLevel+9):
             col_letter = get_column_letter(idx)
             ws_rxn.column_dimensions[col_letter].width = 15
        #
        font_format = Font(name='Times New Roman', bold=True, color=colors.BLACK)
        for cell in ["A1", "B1", "C1", "D1", "F1", "H1"]:
            ws_rxn[cell].font = font_format
        for idx in range(diffThresholdLevel+9, diffThresholdLevel+20):
            col_letter = get_column_letter(idx)
            ws_rxn[col_letter+'1'].font = font_format
        font_format = Font(name='Times New Roman', bold=True, color=colors.BLUE)
        for cell in ["E1", "G1"]:
            ws_rxn[cell].font = font_format
        for idx in range(9, diffThresholdLevel+9):
            col_letter = get_column_letter(idx)
            ws_rxn[col_letter+'1'].font = font_format
        #
        # Create sheet "Compound"
        ws_comp = wb_rxncomp.create_sheet("Compound")
        ws_comp.freeze_panes = "B2"
        # Write title and entries
        for compInfo in compFuseInfoList:
            ws_comp.append(compInfo)
        # Set the format of the sheet
        for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P"]:
            ws_comp.column_dimensions[col_letter].width = 15
        ws_comp.column_dimensions["A"].width = 15
        #
        font_format = Font(name='Times New Roman', bold=True, color=colors.BLACK)
        for cell in ["A1", "B1", "C1", "D1", "E1", "F1", "G1", "H1", "I1", "J1", "K1", "L1", "M1", "N1", "O1", "P1"]:
            ws_comp[cell].font = font_format
        font_format = Font(name='Times New Roman', bold=True, color=colors.BLUE)
        for cell in ["B1", "G1", "H1"]:
            ws_comp[cell].font = font_format
        #
        # Save the excel file
        wb_rxncomp.save(xlsxFile)


    @classmethod
    def _generateInputPrompt(cls, compFuseInfoList, rxnFuseInfoList,
                                  inputprompt_jsonFile = "MetaCyc_InputPrompt.json",
                             ):
        """
        """
        #
        if "temp" not in listdir():
            mkdir("temp")
        inputprompt_jsonFile = path.join("temp", inputprompt_jsonFile)
        #
        comp_inputprompt_list = list()
        for comp_info in compFuseInfoList[1:]:  # Exclude Title
            comp_id = comp_info[0]
            comp_name_list = json.loads(comp_info[1])
            comp_inputprompt_list.append(comp_id)
            for comp_name in comp_name_list:
                comp_inputprompt= "".join([comp_name, "(", comp_id, ")"])
                comp_inputprompt_list.append(comp_inputprompt)
        #
        rxn_inputprompt_list = list()
        for rxn_info in rxnFuseInfoList[1:]:  # Exclude Title
            rxn_id = rxn_info[0]
            rxn_name_list = json.loads(rxn_info[1])
            rxn_inputprompt_list.append(rxn_id)
            for rxn_name in rxn_name_list:
                rxn_inputprompt = "".join([rxn_name, "(", rxn_id, ")"])
                rxn_inputprompt_list.append(rxn_inputprompt)
        #
        inputprompt_metacyc_dict = dict()
        inputprompt_metacyc_dict["Compound"] = comp_inputprompt_list
        inputprompt_metacyc_dict["Reaction"] = rxn_inputprompt_list
        #
        with open(inputprompt_jsonFile, "w") as file_obj:
            file_obj.write(json.dumps(inputprompt_metacyc_dict, indent=2))
        #


    @classmethod
    def _generateDataFile(cls, compFuseInfoList, rxnFuseInfoList,
                               diffThresholdLevel = 10,
                               ):
        """
            Generate .data file to speed up data loading.
        """
        #
        # Direction; Compound List; Compound Set; Compound Pair(diffThreshold); Reference(KEGG, Rhea, MetaCyc, KndPad, EC Number); Status; Coefficient List;
        rxn_distinfo_dict = dict()
        for rxn_info in rxnFuseInfoList[1:]:  # Exclude title
            rxn_id = rxn_info[0]                            # Entry
            rxn_dir = rxn_info[4]                           # Direction
            rxn_dir = rxn_dir.strip('"')
            rxn_comp_list = json.loads(rxn_info[6])         # Compound
            rxn_comp_set = set(rxn_comp_list[0] + rxn_comp_list[1])
            comp_pair_list = rxn_info[8:8+diffThresholdLevel]   # Compound Pair (0.1 - 1.0)
            kegg_rxn_list = json.loads(rxn_info[-5])        # KEGG
            rhea_rxn_list = json.loads(rxn_info[-6])        # Rhea
            metacyc_rxn_list = json.loads(rxn_info[-7])     # MetaCyc
            kndpad_rxn_list = [rxn_id]                      # KndPad
            ec_number_list = json.loads(rxn_info[-9])       # EC Number
            status = rxn_info[-11]                          # Status
            coefficient_list = json.loads(rxn_info[5])      # Coefficient
            reference_rxn_list = [kegg_rxn_list, rhea_rxn_list, metacyc_rxn_list, kndpad_rxn_list, ec_number_list]
            fb_comp_pair_list = list()
            for comp_pair in comp_pair_list:
                if comp_pair == 'null':
                    fb_comp_pair = {'forward':[{}, {}], 'backward':[{}, {}]}
                else:
                    forward_comp_pair = json.loads(comp_pair)
                    # Prepare the compound pairs for pathway retro-search.
                    backward_comp_pair = cls._f2b(forward_comp_pair)
                    fb_comp_pair = {'forward':forward_comp_pair, 'backward':backward_comp_pair}
                fb_comp_pair_list.append(fb_comp_pair)
            rxn_distinfo_dict[rxn_id] = [rxn_dir, rxn_comp_list, rxn_comp_set, fb_comp_pair_list, reference_rxn_list, status, coefficient_list]
        #
        # Name; Branching Factor; Reactions; Reference(KEGG, ChEBI, MetaCyc, KndPad); SMILES;
        comp_distinfo_dict = dict()
        for comp_info in compFuseInfoList[1:]:  # Exclude title
            comp_id = comp_info[0]                          # Entry
            comp_name_list = json.loads(comp_info[1])       # Names
            comp_labf_list = json.loads(comp_info[6])       # Branching Factor
            comp_rxn_list = json.loads(comp_info[7])        # Reactions
            kegg_comp_list = json.loads(comp_info[-5])      # KEGG
            chebi_comp_list = json.loads(comp_info[-6])     # ChEBI
            metacyc_comp_list = json.loads(comp_info[-7])   # MetaCyc
            kndpad_comp_list = [comp_id]                    # KndPad
            smiles_comp = comp_info[5]                      # SMILES
            reference_comp_list = [kegg_comp_list, chebi_comp_list, metacyc_comp_list, kndpad_comp_list]
            comp_distinfo_dict[comp_id] = [comp_name_list, comp_labf_list, comp_rxn_list, reference_comp_list, smiles_comp]
        #
        #
        data_file = "MetaCyc_Dmn.data"
        with open(path.join("temp", data_file), "wb") as file_obj:
            pickle.dump([rxn_distinfo_dict, comp_distinfo_dict], file_obj)
        #
        return None


    @classmethod
    def _f2b(cls, forward_comp_pair):
        """
            Prepare the compound pairs for pathway retro-search.
        """
        rpf_comp_pair_dict = forward_comp_pair[0]
        prf_comp_pair_dict = forward_comp_pair[1]
        #
        rpb_comp_pair_dict = dict()
        for comp_id_list in rpf_comp_pair_dict.values():
            for comp_id in comp_id_list:
                rpb_comp_pair_dict[comp_id] = list()
        prb_comp_pair_dict = dict()
        for comp_id_list in prf_comp_pair_dict.values():
            for comp_id in comp_id_list:
                prb_comp_pair_dict[comp_id] = list()
        #
        for key, comp_id_list in rpf_comp_pair_dict.items():
            for comp_id in comp_id_list:
                rpb_comp_pair_dict[comp_id].append(key)
        for key, comp_id_list in prf_comp_pair_dict.items():
            for comp_id in comp_id_list:
                prb_comp_pair_dict[comp_id].append(key)
        #
        backward_comp_pair = [rpb_comp_pair_dict, prb_comp_pair_dict]
        return backward_comp_pair


    @classmethod
    def loadRxnCompPair(cls, dirSourceData = "../sourcedata/",
                             xlsxFile = "MetaCyc_Pathway_Search_Ori.xlsx",
                             diffThreshold = 0.1,
                             isUpdate = True,
                             diffThresholdLevel = 10,
                        ):
        """
            Load compound pairs of every reaction from Metacyc.
        """
        #
        data_file = "MetaCyc_Dmn.data"
        xlsxFile = path.join(dirSourceData, xlsxFile)
        #
        if isUpdate or (data_file not in listdir("temp")):
            reac_info_list = InputMeta._getSheetInfo(xlsxFile, sheetName="Reaction")
            comp_info_list = InputMeta._getSheetInfo(xlsxFile, sheetName="Compound")
            #
            # Direction; Compound List; Compound Set; Compound Pair(diffThreshold); Reference(KEGG, Rhea, MetaCyc, KndPad, EC Number); Status; Coefficient List;
            rxn_distinfo_dict = dict()
            for rxn_info in reac_info_list[1:]:  # Exclude title
                rxn_id = rxn_info[0]                            # Entry
                rxn_dir = rxn_info[4]                           # Direction
                rxn_dir = rxn_dir.strip('"')
                rxn_comp_list = json.loads(rxn_info[6])         # Compound
                rxn_comp_set = set(rxn_comp_list[0] + rxn_comp_list[1])
                comp_pair_list = rxn_info[8:8+diffThresholdLevel]   # Compound Pair (0.1 - 1.0)
                kegg_rxn_list = json.loads(rxn_info[-5])        # KEGG
                rhea_rxn_list = json.loads(rxn_info[-6])        # Rhea
                metacyc_rxn_list = json.loads(rxn_info[-7])     # MetaCyc
                kndpad_rxn_list = [rxn_id]                      # KndPad
                ec_number_list = json.loads(rxn_info[-9])       # EC Number
                status = rxn_info[-11]                          # Status
                coefficient_list = json.loads(rxn_info[5])      # Coefficient
                reference_rxn_list = [kegg_rxn_list, rhea_rxn_list, metacyc_rxn_list, kndpad_rxn_list, ec_number_list]
                fb_comp_pair_list = list()
                for comp_pair in comp_pair_list:
                    if comp_pair == 'null':
                        fb_comp_pair = {'forward':[{}, {}], 'backward':[{}, {}]}
                    else:
                        forward_comp_pair = json.loads(comp_pair)
                        # Prepare the compound pairs for pathway retro-search.
                        backward_comp_pair = cls._f2b(forward_comp_pair)
                        fb_comp_pair = {'forward':forward_comp_pair, 'backward':backward_comp_pair}
                    fb_comp_pair_list.append(fb_comp_pair)
                rxn_distinfo_dict[rxn_id] = [rxn_dir, rxn_comp_list, rxn_comp_set, fb_comp_pair_list, reference_rxn_list, status, coefficient_list]
            #
            # Name; Branching Factor; Reactions; Reference(KEGG, ChEBI, MetaCyc, KndPad); SMILES;
            comp_distinfo_dict = dict()
            for comp_info in comp_info_list[1:]:  # Exclude title
                comp_id = comp_info[0]                          # Entry
                comp_name_list = json.loads(comp_info[1])       # Names
                comp_labf_list = json.loads(comp_info[6])       # Branching Factor
                comp_rxn_list = json.loads(comp_info[7])        # Reactions
                kegg_comp_list = json.loads(comp_info[-5])      # KEGG
                chebi_comp_list = json.loads(comp_info[-6])     # ChEBI
                metacyc_comp_list = json.loads(comp_info[-7])   # MetaCyc
                kndpad_comp_list = [comp_id]                    # KndPad
                smiles_comp = comp_info[5]                      # SMILES
                reference_comp_list = [kegg_comp_list, chebi_comp_list, metacyc_comp_list, kndpad_comp_list]
                comp_distinfo_dict[comp_id] = [comp_name_list, comp_labf_list, comp_rxn_list, reference_comp_list, smiles_comp]
            #
            # Updating...
            with open(path.join("temp", data_file), "wb") as file_obj:
                pickle.dump([rxn_distinfo_dict, comp_distinfo_dict], file_obj)
        else:
            # Loading...
            with open(path.join("temp", data_file), "rb") as file_obj:
                [rxn_distinfo_dict, comp_distinfo_dict] = pickle.load(file_obj)
        #
        idx = int(10*diffThreshold) - 1
        #
        # Direction; Compound List; Compound Set; Compound Pair(diffThreshold); Reaction Reference List; Status; Coefficient;
        rxnDistInfoDict = dict()
        for key, value in rxn_distinfo_dict.items():
            rxnDistInfoDict[key] = [value[0], value[1], value[2], value[3][idx], value[4], value[5], value[6]]
        #
        # Name; Branching Factor; Reactions; Compound Reference List; SMILES;
        compDistInfoDict = dict()
        for key, value in comp_distinfo_dict.items():
            compDistInfoDict[key] = [value[0][0], value[1][idx], value[2], value[3], value[4]]
        #
        # Direction; Compound List; Compound Set; Compound Pair(diffThreshold); Reaction Reference List; Status; Coefficient;
        # Name; Branching Factor; Reactions; Compound Reference List; SMILES;
        return [rxnDistInfoDict, compDistInfoDict]


    @classmethod
    def _getOrgCompRxnId(cls, org_model_location = "../sourcedata/models/",
                              organism = "eco",
                              json_model_file_dict = dict(),
                              dataBase = "MetaCyc",
                              xref_metacyc2kndpad_json = "datas/MetaCyc_Xref_KndPad.json",
                              xref_metacyc2bigg_json = "datas/KndPad_Xref_BiGG.json",
                              ):
        """
            Get the information of the selected organism from bigg.
            And we try to map compounds and reactions in MetaCyc(KndPad) to the metabolites and metabolic reactions in cytosol
        of this organism.
        """
        #
        orgCompRxnId_dict = dict()
        if organism == "":
            return orgCompRxnId_dict
        #
        with open(xref_metacyc2kndpad_json, "r") as file_obj:
            metacyc_xref_kndpad_dict = json.load(file_obj)
        #
        # "bsu": "bsu_iYO844.json",
        # "eco": "eco_iJO1366.json",
        # "kpn": "kpn_iYL1228.json",
        # "ppu": "ppu_iJN1462.json",
        # "sce": "sce_iMM904.json",
        # "syz": "syz_iSynCJ816.json",
        # ...
        #
        org_json_file = json_model_file_dict[organism]
        org_json_file = path.join(org_model_location, org_json_file)
        with open(org_json_file, "r") as file_obj:
            model_dict = json.load(file_obj)
        #
        metabolite_list = model_dict['metabolites']
        reaction_list = model_dict['reactions']
        #
        #
        # Compound from model
        met_id_set = set()
        compartment_set = set()
        comp_metacyc2model_dict = dict()
        for met_info_dict in metabolite_list:
            met_id = met_info_dict['id']
            met_id_set.add(met_id)
            compartment = met_info_dict['compartment']
            compartment_set.add(compartment)
            met_annotation_dict = met_info_dict['annotation']
            try:
                biocyc_comp_list = met_annotation_dict['biocyc']
                for biocyc_comp_id in biocyc_comp_list:
                    # biocyc --> metacyc --> kndpad
                    metacyc_comp_id = biocyc_comp_id.replace("META:", "")
                    try:
                        kndpad_comp_id = metacyc_xref_kndpad_dict["Compound"][metacyc_comp_id]
                    except KeyError:
                        continue
                    try:
                        if comp_metacyc2model_dict[kndpad_comp_id].endswith('_c'):  # cytosol
                            pass
                        else:
                            comp_metacyc2model_dict[kndpad_comp_id] = met_id
                    # new item
                    except KeyError:
                        comp_metacyc2model_dict[kndpad_comp_id] = met_id
            except KeyError:
                pass
        #
        # print("compartments of ", organism, ":")
        # print(compartment_set)
        # compartments of  bsu : {'e', 'c'}
        # compartments of  eco : {'e', 'c', 'p'}
        # compartments of  sce : {'x', 'v', 'm', 'g', 'e', 'r', 'n', 'c'}
        #
        # c: cytosol / cytoplasm
        # m: mitochondria / mitochondrion
        # n: nucleus
        # r: endoplasmic reticulum
        # g: golgi apparatus / golgi
        # v: vacuole
        # x: peroxisome / glyoxysome
        # p: periplasm / periplasmic space
        # e: extracellular space / extra-organism / extraorganism / extracellular
        #
        # unity...
        unify_compartment_list = ['c', 'm', 'n', 'r', 'g', 'v', 'x', 'p', 'e']
        compartment_list = list()
        for compartment in unify_compartment_list:
            if compartment in compartment_set:
                compartment_list.append(compartment)
        #
        #
        # Reaction from model
        rxn_id_set = set()
        reac_metacyc2model_dict = dict()
        for rxn_info_dict in reaction_list:
            rxn_id = rxn_info_dict['id']
            rxn_id_set.add(rxn_id)
            rxn_annotation_dict = rxn_info_dict['annotation']
            try:
                biocyc_reac_list = rxn_annotation_dict['biocyc']
                for biocyc_reac_id in biocyc_reac_list:
                    # biocyc --> metacyc --> kndpad
                    metacyc_reac_id = biocyc_reac_id.replace("META:", "")
                    try:
                        kndpad_reac_id = metacyc_xref_kndpad_dict["Reaction"][metacyc_reac_id]
                    except KeyError:
                        continue
                    reac_metacyc2model_dict[kndpad_reac_id] = rxn_id
            except KeyError:
                pass
        #
        #
        # Compound from reference
        # Reaction from reference
        if path.exists(xref_metacyc2bigg_json):
            with open(xref_metacyc2bigg_json, "r") as file_obj:
                xref_metacyc2bigg_list_dict = json.load(file_obj)
            xref_metacyc2bigg_comp_list_dict = xref_metacyc2bigg_list_dict['Compound']
            xref_metacyc2bigg_rxn_list_dict = xref_metacyc2bigg_list_dict['Reaction']
        else:
            xref_metacyc2bigg_comp_list_dict = dict()
            xref_metacyc2bigg_rxn_list_dict = dict()
        #
        # Compound from reference
        for kndpad_comp_id, bigg_met_id_list in xref_metacyc2bigg_comp_list_dict.items():
            # without reference
            if (len(bigg_met_id_list) == 0) or (not kndpad_comp_id.endswith('-m')):
                continue
            try:
                # existence
                if comp_metacyc2model_dict[kndpad_comp_id].endswith('_c'):  # cytosol
                    pass
                else:
                    for bigg_met_id in bigg_met_id_list:
                        met_id = "".join([bigg_met_id, "_c"])
                        if met_id in met_id_set:
                            comp_metacyc2model_dict[kndpad_comp_id] = met_id     # other compartment --> cytosol
            except KeyError:
                # bigg id
                for bigg_met_id in bigg_met_id_list:
                    # compartment
                    for compartment in compartment_list:
                        met_id = "".join([bigg_met_id, "_", compartment])
                        if met_id in met_id_set:
                            comp_metacyc2model_dict[kndpad_comp_id] = met_id
                            break
                    try:
                        if comp_metacyc2model_dict[kndpad_comp_id].endswith('_c'):  # cytosol
                            break
                    except KeyError:
                        pass
        #
        # Reaction from reference
        for kndpad_reac_id, bigg_rxn_id_list in xref_metacyc2bigg_rxn_list_dict.items():
            # without reference
            if (len(bigg_rxn_id_list) == 0) or (not kndpad_reac_id.endswith('-m') ):
                continue
            try:
                # existence
                rxn_id = reac_metacyc2model_dict[kndpad_reac_id]
            except KeyError:
                for bigg_rxn_id in bigg_rxn_id_list:
                    rxn_id = bigg_rxn_id
                    if rxn_id in rxn_id_set:
                        reac_metacyc2model_dict[kndpad_reac_id] = rxn_id
                        break
        #
        #
        # Compound(bsu)/Reaction(bsu), Compound(eco)/Reaction(eco) or Compound(sce)/Reaction(sce).
        comp_key = "".join(["Compound", "(", organism, ")"])
        orgCompRxnId_dict[comp_key] = comp_metacyc2model_dict
        reac_key = "".join(["Reaction", "(", organism, ")"])
        orgCompRxnId_dict[reac_key] = reac_metacyc2model_dict
        #
        return orgCompRxnId_dict


    @classmethod
    def demoFunc(cls,):
        """ """
        pass



if __name__ == "__main__":
    """ """
    from os import getcwd, chdir
    chdir("../../")
    print("Current Path: ", getcwd())
    [compDistInfoList, rxnDistInfoList] = InputMetaCyc._getMetaCycCompRxnDistInfo()





