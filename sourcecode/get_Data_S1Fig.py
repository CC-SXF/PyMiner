# -*- coding: utf-8 -*-
"""
Created on Thu Mar  3 23:04:52 2022

@author: CC-SXF
"""

from openpyxl import load_workbook

dir_metacyc_pathway = './datas/MetaCyc_Validation_Pathway.xlsx'
dir_kndpad = '../sourcedata/KndPad_Ori_Cleaned_Half.xlsx'


def getSheetAllRows(sheetName):
    """
        Read all rows in 'sheetName'.
    """
    sheetAllRows = []
    for row in sheetName.rows:
        sheetOneRow = []
        for cell in row:
            sheetOneRow.append(cell.value)
        sheetAllRows.append(sheetOneRow)
    return sheetAllRows


wb_pwy = load_workbook(dir_metacyc_pathway)
wb_pwy.guess_types = True
ws_pwy = wb_pwy["Pathway"]
pwy_info_list = getSheetAllRows(ws_pwy)

wb_kndpad = load_workbook(dir_kndpad)
wb_kndpad.guess_types = True
ws_kndpad_comp = wb_kndpad['Compound']
comp_info_list = getSheetAllRows(ws_kndpad_comp)


# S1A Fig
comp_rxnNum = dict()
for comp_info in comp_info_list[1:]:  # Exclude title
    comp_rxnNum[comp_info[0]] = int(comp_info[6])

# '"Biosynthesis -> '
# '"Degradation/Utilization/Assimilation ->'

# 'Secondary Metabolite Biosynthesis'
# 'Secondary Metabolite Degradation'

sec_met_set = set()
pri_met_set = set()
for pwy_info in pwy_info_list:
    if 'Secondary Metabolite' in pwy_info[3]:
        if 'Secondary Metabolite Biosynthesis' in pwy_info[3]:
            sec_met_set.add( pwy_info[2][-11:] )
        if 'Secondary Metabolite Degradation' in pwy_info[3]:
            sec_met_set.add( pwy_info[2][:11] )
    else:
        if '"Biosynthesis -> '  in pwy_info[3]:
            pri_met_set.add( pwy_info[2][-11:] )
        if '"Degradation/Utilization/Assimilation ->' in pwy_info[3]:
            pri_met_set.add( pwy_info[2][:11] )

intersection_met = (sec_met_set & pri_met_set) # 交集
sec_met_list = sorted(sec_met_set-intersection_met)
pri_met_list = sorted(pri_met_set-intersection_met)


sec_met_rxnNum = {met_id:comp_rxnNum[met_id] for met_id in sec_met_list}
pri_met_rxnNum = {met_id:comp_rxnNum[met_id] for met_id in pri_met_list}

print('')
print('Number of Secondary Metabolite', len(sec_met_rxnNum))
print("Average Reaction Number of Secondary Metabolite:", round(sum(sec_met_rxnNum.values())/len(sec_met_rxnNum), 3))
print('Number of Non-Secondary Metabolite:', len(pri_met_rxnNum))
print("Average Reaction Number of Non-Secondary Metabolite:", round(sum(pri_met_rxnNum.values())/len(pri_met_rxnNum), 3))
print('')



# S1B Fig
# primary to secondary or secondary to primary
# 490
dir_pssp_num = 0
dir_pssp_dict = {}

for pwy_info in pwy_info_list[1:]:  #exclude title
    pwy_id = pwy_info[0]
    pwy_str = pwy_info[2]
    pwy_metrxn_list = pwy_str.split('-->')
    (sMet, tMet) = (pwy_metrxn_list[0], pwy_metrxn_list[-1])

    # primary to secondary
    if (sMet in pri_met_list) and (tMet in sec_met_list):
        dir_pssp_dict[pwy_id] = 1
        dir_pssp_num += 1

    # secondary to primary
    elif (sMet in sec_met_list) and (tMet in pri_met_list):
        dir_pssp_dict[pwy_id] = -1
        dir_pssp_num += 1

    else:
        dir_pssp_dict[pwy_id] = 0

print("Primary metabolite to secondary metabolite or secondary metabolite to primary metabolite")
print("number(p2s_s2p):", dir_pssp_num)
print('')


if __name__ == "__main__":
    """
    """
    pass





